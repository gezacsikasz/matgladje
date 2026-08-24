#!/usr/bin/env python3
"""Konverterar data/matgladje_skafferi.xlsx till prisma/seed-data/skafferi.json.

Körs manuellt (kräver `pip install openpyxl`) när källexcel-filen ändras.
Seed-scriptet (prisma/seed.ts) läser bara den genererade JSON-filen, inte
Excel direkt, för att undvika en xlsx-parsningsberoende i själva Next.js-appen.
"""
import json
import openpyxl

SRC = "data/matgladje_skafferi.xlsx"
OUT = "prisma/seed-data/skafferi.json"

wb = openpyxl.load_workbook(SRC, data_only=True)

# --- Skafferi: kategorier + produkter -------------------------------------
skafferi_rows = list(wb["Skafferi"].iter_rows(min_row=2, values_only=True))
categories = []
seen_categories = {}
products = []
for row in skafferi_rows:
    if not row or not row[0]:
        continue
    (pid, kategori, namn, enhet, butik, ekobutik, reko, malpris,
     vs_butik, vs_reko, sakerhet, kommentar) = row
    if kategori not in seen_categories:
        seen_categories[kategori] = len(categories)
        categories.append({"namn": kategori, "sortOrder": len(categories)})
    products.append({
        "id": pid,
        "kategori": kategori,
        "namn": namn,
        "enhet": enhet,
        "butikReferens": butik,
        "ekobutikReferens": ekobutik,
        "rekoReferens": reko,
        "malpris": malpris,
        "sakerhet": sakerhet,
        "kommentar": kommentar,
        "sourcingStatus": None,
        "sourcingNote": None,
    })

# --- Behovsmodell: kr/KE/manad per kategori --------------------------------
behov_rows = list(wb["Behovsmodell (75p)"].iter_rows(values_only=True))
total_ke = None
for row in behov_rows:
    if row and row[0] == "Summa":
        total_ke = row[3]
        break
assert total_ke, "Kunde inte hitta total KE-summa i Behovsmodell-bladet"

behov_by_category = {}
in_section3 = False
for row in behov_rows:
    if row and row[0] == "3. Summering per kategori (kohortens årsvärde)":
        in_section3 = True
        continue
    if in_section3:
        if not row or row[0] in (None, "Kategori"):
            if row and row[0] == "TOTALT":
                break
            continue
        kategori, arsvarde, manadsvarde, veckovarde = row[0], row[1], row[2], row[3]
        behov_by_category[kategori] = round(manadsvarde / total_ke, 4)

for cat in categories:
    cat["krPerKePerManad"] = behov_by_category.get(cat["namn"])

# --- Producenter: en rad per produkt, grupperas till unika producenter -----
# "-" i namnkolumnen betyder att ingen specifik producentkandidat är
# identifierad för produkten än (status/källa förklarar varför/vad som
# återstår att göra) — det ska INTE tolkas som en producent som heter "-".
products_by_id = {p["id"]: p for p in products}
prod_rows = list(wb["Producenter"].iter_rows(min_row=2, values_only=True))
producers_by_name = {}
producer_order = []
for row in prod_rows:
    if not row or not row[0]:
        continue
    (pid, _produkt_auto, producent_namn, region, webb, leveranssatt,
     status, kalla) = row
    if not producent_namn or producent_namn == "-":
        if pid in products_by_id:
            products_by_id[pid]["sourcingStatus"] = status
            products_by_id[pid]["sourcingNote"] = kalla
        continue
    if producent_namn not in producers_by_name:
        producers_by_name[producent_namn] = {
            "namn": producent_namn,
            "region": region,
            "egenWebb": None if webb in (None, "-") else webb,
            "sourcingLeveranssatt": leveranssatt,
            "sourcingKalla": kalla,
            "produkter": [],
        }
        producer_order.append(producent_namn)
    producers_by_name[producent_namn]["produkter"].append(pid)

producers = [producers_by_name[name] for name in producer_order]

out = {
    "categories": categories,
    "products": products,
    "producers": producers,
}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f"Skrev {OUT}: {len(categories)} kategorier, {len(products)} produkter, "
      f"{len(producers)} unika producentkandidater.")
